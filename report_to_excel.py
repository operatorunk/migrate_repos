import requests
import urllib3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

BITBUCKET_URL = "https://bitbucket.tuempresa.com"
USERNAME = "TU_USUARIO"
TOKEN = "TU_TOKEN"
PROJECT_KEY = "TU_PROJECT_KEY"

STALE_BRANCH_DAYS = 180
REPORT_FILE = "Bitbucket_Migration_Report.xlsx"


def get_paginated(url):
    results = []
    start = 0
    separator = "&" if "?" in url else "?"

    while True:
        paged_url = f"{url}{separator}start={start}"

        response = requests.get(
            paged_url,
            auth=(USERNAME, TOKEN),
            verify=False
        )

        response.raise_for_status()
        data = response.json()

        results.extend(data.get("values", []))

        if data.get("isLastPage", True):
            break

        start = data["nextPageStart"]

    return results


def format_timestamp(timestamp_ms):
    if not timestamp_ms:
        return "N/A"

    return datetime.fromtimestamp(timestamp_ms / 1000).strftime("%Y-%m-%d %H:%M:%S")


def timestamp_to_datetime(timestamp_ms):
    if not timestamp_ms:
        return None

    return datetime.fromtimestamp(timestamp_ms / 1000)


def get_repositories():
    url = f"{BITBUCKET_URL}/rest/api/1.0/projects/{PROJECT_KEY}/repos"
    return get_paginated(url)


def get_open_pull_requests(repo_slug):
    url = f"{BITBUCKET_URL}/rest/api/1.0/projects/{PROJECT_KEY}/repos/{repo_slug}/pull-requests?state=OPEN"
    return get_paginated(url)


def get_branches(repo_slug):
    url = f"{BITBUCKET_URL}/rest/api/1.0/projects/{PROJECT_KEY}/repos/{repo_slug}/branches"
    return get_paginated(url)


def get_tags(repo_slug):
    url = f"{BITBUCKET_URL}/rest/api/1.0/projects/{PROJECT_KEY}/repos/{repo_slug}/tags"
    return get_paginated(url)


def get_commit(repo_slug, commit_id):
    url = f"{BITBUCKET_URL}/rest/api/1.0/projects/{PROJECT_KEY}/repos/{repo_slug}/commits/{commit_id}"

    response = requests.get(
        url,
        auth=(USERNAME, TOKEN),
        verify=False
    )

    response.raise_for_status()
    return response.json()


def get_default_branch_from_branches(branches):
    for branch in branches:
        if branch.get("isDefault") is True:
            return branch.get("displayId", "N/A")

    return "N/A"


def get_repo_size(repo):
    return repo.get("size", "N/A")


def is_stale_branch(commit_datetime):
    if not commit_datetime:
        return False

    age_days = (datetime.now() - commit_datetime).days
    return age_days > STALE_BRANCH_DAYS


def auto_fit_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 45)


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    auto_fit_columns(ws)


def create_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)

    for row in rows:
        ws.append(row)

    style_sheet(ws)
    return ws


def build_excel(repository_rows, branch_rows, pr_rows, tag_rows, summary_rows):
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    create_sheet(
        wb,
        "Repositories",
        [
            "Project Key",
            "Repo Name",
            "Repo Slug",
            "Default Branch",
            "Repo Size",
            "Branch Count",
            "Tag Count",
            "Open PR Count",
            "Latest Activity",
        ],
        repository_rows,
    )

    create_sheet(
        wb,
        "Branches",
        [
            "Project Key",
            "Repo Name",
            "Repo Slug",
            "Branch Name",
            "Latest Commit",
            "Author",
            "Email",
            "Commit Date",
            "Status",
        ],
        branch_rows,
    )

    create_sheet(
        wb,
        "Pull Requests",
        [
            "Project Key",
            "Repo Name",
            "Repo Slug",
            "PR ID",
            "Title",
            "Author",
            "Source Branch",
            "Target Branch",
        ],
        pr_rows,
    )

    create_sheet(
        wb,
        "Tags",
        [
            "Project Key",
            "Repo Name",
            "Repo Slug",
            "Tag Name",
            "Latest Commit",
        ],
        tag_rows,
    )

    create_sheet(
        wb,
        "Migration Summary",
        ["Metric", "Value"],
        summary_rows,
    )

    create_sheet(
        wb,
        "Users Manual",
        [
            "Project Key",
            "Repo Name",
            "User",
            "Permission",
            "Source",
            "Notes",
        ],
        [],
    )

    create_sheet(
        wb,
        "Groups Manual",
        [
            "Project Key",
            "Repo Name",
            "Group",
            "Permission",
            "Members",
            "Source",
            "Notes",
        ],
        [],
    )

    wb.save(REPORT_FILE)


def main():
    repos = get_repositories()

    repository_rows = []
    branch_rows = []
    pr_rows = []
    tag_rows = []

    total_branches = 0
    total_active_branches = 0
    total_stale_branches = 0
    total_open_prs = 0
    total_tags = 0

    for repo in repos:
        repo_slug = repo["slug"]
        repo_name = repo["name"]
        repo_size = get_repo_size(repo)

        open_prs = get_open_pull_requests(repo_slug)
        branches = get_branches(repo_slug)
        default_branch = get_default_branch_from_branches(branches)
        tags = get_tags(repo_slug)

        latest_repo_activity = None

        print("=" * 80)
        print(f"Repo: {repo_name} ({repo_slug})")
        print(f"Default Branch: {default_branch}")
        print(f"Repo Size: {repo_size}")
        print(f"Open PRs: {len(open_prs)}")
        print(f"Branches: {len(branches)}")
        print(f"Tags: {len(tags)}")

        total_open_prs += len(open_prs)
        total_branches += len(branches)
        total_tags += len(tags)

        for pr in open_prs:
            author = pr.get("author", {}).get("user", {}).get("displayName", "N/A")
            from_branch = pr.get("fromRef", {}).get("displayId", "N/A")
            to_branch = pr.get("toRef", {}).get("displayId", "N/A")

            pr_rows.append([
                PROJECT_KEY,
                repo_name,
                repo_slug,
                pr.get("id", ""),
                pr.get("title", ""),
                author,
                from_branch,
                to_branch,
            ])

        for branch in branches:
            branch_name = branch.get("displayId", "N/A")
            latest_commit = branch.get("latestCommit", "N/A")

            commit = get_commit(repo_slug, latest_commit)

            author = commit.get("author", {}).get("name", "N/A")
            author_email = commit.get("author", {}).get("emailAddress", "N/A")
            commit_timestamp = commit.get("authorTimestamp")
            commit_date = format_timestamp(commit_timestamp)
            commit_datetime = timestamp_to_datetime(commit_timestamp)

            if commit_datetime and (
                latest_repo_activity is None or commit_datetime > latest_repo_activity
            ):
                latest_repo_activity = commit_datetime

            branch_status = "STALE" if is_stale_branch(commit_datetime) else "ACTIVE"

            if branch_status == "STALE":
                total_stale_branches += 1
            else:
                total_active_branches += 1

            branch_rows.append([
                PROJECT_KEY,
                repo_name,
                repo_slug,
                branch_name,
                latest_commit,
                author,
                author_email,
                commit_date,
                branch_status,
            ])

        latest_repo_activity_text = (
            latest_repo_activity.strftime("%Y-%m-%d %H:%M:%S")
            if latest_repo_activity
            else "N/A"
        )

        repository_rows.append([
            PROJECT_KEY,
            repo_name,
            repo_slug,
            default_branch,
            repo_size,
            len(branches),
            len(tags),
            len(open_prs),
            latest_repo_activity_text,
        ])

        for tag in tags:
            tag_rows.append([
                PROJECT_KEY,
                repo_name,
                repo_slug,
                tag.get("displayId", "N/A"),
                tag.get("latestCommit", "N/A"),
            ])

    summary_rows = [
        ["Project Key", PROJECT_KEY],
        ["Repositories", len(repos)],
        ["Branches", total_branches],
        ["Active Branches", total_active_branches],
        ["Stale Branches", total_stale_branches],
        ["Open Pull Requests", total_open_prs],
        ["Tags", total_tags],
        ["Stale Branch Threshold Days", STALE_BRANCH_DAYS],
        ["Users/Groups", "Pending manual input from Bitbucket interface"],
    ]

    build_excel(
        repository_rows,
        branch_rows,
        pr_rows,
        tag_rows,
        summary_rows,
    )

    print("=" * 80)
    print(f"Excel report generated: {REPORT_FILE}")


if __name__ == "__main__":
    main()
